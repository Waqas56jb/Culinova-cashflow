# Parses every data sheet of CULINOVA_CashFlow_CFO.xlsx into clean JSON
# for import into the database. Run from repo root:
#   python server/db/parse-excel.py
import sys, io, json
from datetime import datetime, timedelta, date
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

wb = openpyxl.load_workbook('CULINOVA_CashFlow_CFO.xlsx', data_only=True)

def to_date(v):
    if v in (None, ''):
        return None
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, date):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, (int, float)):
        if 40000 < v < 60000:  # Excel serial date
            return (datetime(1899, 12, 30) + timedelta(days=float(v))).strftime('%Y-%m-%d')
        return None
    s = str(v).strip()
    try:
        return datetime.fromisoformat(s).strftime('%Y-%m-%d')
    except Exception:
        return None

def to_bool(v):
    if v is None:
        return False
    return str(v).strip().lower() in ('yes', 'true', 'y', '1', 'paid', 'confirmed', 'done')

def to_num(v):
    if v in (None, ''):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(',', '').replace('SAR', '').strip()
    if s in ('', '-', '—'):
        return None
    try:
        return float(s)
    except Exception:
        return None

def to_str(v):
    if v is None:
        return None
    s = str(v).replace('\xa0', ' ').strip()
    return s if s != '' else None

def cell(ws, r, c):
    return ws.cell(r, c).value

def rows_of(name, ncols, key_col=1):
    ws = wb[name]
    out = []
    for r in range(2, ws.max_row + 1):
        vals = [cell(ws, r, c) for c in range(1, ncols + 1)]
        if all(v in (None, '') for v in vals):
            continue
        out.append((r, vals))
    return out

data = {}

# ---------- Collections ----------
data['collections'] = []
for r, v in rows_of('Collections', 12):
    if to_str(v[0]) is None and to_num(v[2]) is None:
        continue
    data['collections'].append({
        'project': to_str(v[0]), 'customer': to_str(v[1]), 'amount': to_num(v[2]),
        'expected_date': to_date(v[3]), 'probability_pct': to_num(v[4]),
        'probability_status': to_str(v[5]), 'invoice_ref': to_str(v[6]),
        'confirmed': to_bool(v[7]), 'actual_collection_date': to_date(v[8]),
        'owner': to_str(v[10]), 'notes': to_str(v[11]),
    })

# ---------- Payments ----------
data['payments'] = []
for r, v in rows_of('Payments', 12):
    if to_str(v[0]) is None and to_str(v[1]) is None and to_num(v[2]) is None:
        continue
    data['payments'].append({
        'category': to_str(v[0]), 'supplier': to_str(v[1]), 'amount': to_num(v[2]),
        'due_date': to_date(v[3]), 'priority': to_str(v[4]), 'qty': to_num(v[5]),
        'project_link': to_str(v[6]), 'can_delay': to_bool(v[7]), 'paid': to_bool(v[8]),
        'actual_payment_date': to_date(v[9]), 'owner': to_str(v[10]), 'notes': to_str(v[11]),
    })

# ---------- Projects Summary ----------
data['projects'] = []
for r, v in rows_of('Projects Summary', 10):
    if to_str(v[0]) is None:
        continue
    data['projects'].append({
        'name': to_str(v[0]), 'contract_value': to_num(v[1]), 'collected_to_date': to_num(v[2]),
        # Client standard expected margin = 30% (overrides the 0.2 in the Excel sheet)
        'gross_profit_pct': 0.30, 'status': to_str(v[6]), 'next_billing': to_str(v[7]),
        'expected_collection_date': to_date(v[8]), 'notes': to_str(v[9]),
    })

# ---------- Inventory ----------
data['inventory'] = []
for r, v in rows_of('Inventory & Stock', 14):
    if to_str(v[0]) is None and to_str(v[2]) is None:
        continue
    data['inventory'].append({
        'category': to_str(v[0]), 'brand_supplier': to_str(v[1]), 'item': to_str(v[2]),
        'qty': to_num(v[3]), 'unit_cost': to_num(v[4]), 'sell_price': to_num(v[6]),
        'stock_location': to_str(v[9]), 'funding_source': to_str(v[10]),
        'linked_project': to_str(v[11]), 'stock_risk': to_str(v[12]), 'notes': to_str(v[13]),
    })

# ---------- Supplier Ledger ----------
# po stored as (p.o + vat) / 1.15 so that derived po+vat & closing match Excel exactly.
data['supplier_ledger'] = []
for r, v in rows_of('Supplier Ledger', 8):
    supplier = to_str(v[0])
    if supplier is None:
        continue
    povat = to_num(v[2])
    po_raw = to_num(v[1])
    po = (povat / 1.15) if povat is not None else po_raw
    data['supplier_ledger'].append({
        'supplier': supplier, 'po': po, 'invoiced_amount': to_num(v[3]),
        'paid_amount': to_num(v[4]), 'negotiation_action': to_str(v[7]),
    })

# ---------- Customer Ledger ----------
data['customer_ledger'] = []
for r, v in rows_of('Customer Ledger', 11):
    if to_str(v[0]) is None and to_num(v[2]) is None:
        continue
    data['customer_ledger'].append({
        'project_name': to_str(v[0]), 'project_id': to_str(v[1]), 'so': to_num(v[2]),
        'invoiced_amount': to_num(v[4]), 'paid_amount': to_num(v[5]),
        'project_status': to_str(v[9]), 'notes': to_str(v[10]),
    })

# ---------- AR Aging ----------
data['ar_aging'] = []
for r, v in rows_of('AR Aging', 9):
    has_amt = any(to_num(v[c]) for c in range(1, 6))
    if to_str(v[0]) is None and not has_amt:
        continue
    data['ar_aging'].append({
        'customer': to_str(v[0]) or '(unnamed)', 'current_amt': to_num(v[1]),
        'd1_30': to_num(v[2]), 'd31_60': to_num(v[3]), 'd61_90': to_num(v[4]),
        'd90_plus': to_num(v[5]), 'collection_owner': to_str(v[7]),
        'action_required': to_str(v[8]),
    })

json.dump(data, open('server/db/excel_import.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=0)
for k, v in data.items():
    print(f'{k:18} {len(v)} rows')
